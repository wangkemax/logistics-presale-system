"""Quotation Excel generator.

Produces a professionally formatted .xlsx file with:
- Sheet 1: 报价汇总 (Summary + financial indicators)
- Sheet 2: 成本明细 (5-category cost breakdown, 3-year)
- Sheet 3: 人力明细 (Labor cost details)
- Sheet 4: 价格方案 (Pricing scheme)
"""

import io
from datetime import datetime

from openpyxl import Workbook
from openpyxl.styles import (
    Font, PatternFill, Alignment, Border, Side, numbers,
)
from openpyxl.utils import get_column_letter


# ── Style constants ──

HEADER_FILL = PatternFill(start_color="1F3864", end_color="1F3864", fill_type="solid")
HEADER_FONT = Font(name="微软雅黑", bold=True, color="FFFFFF", size=11)
SUBHEADER_FILL = PatternFill(start_color="D6E4F0", end_color="D6E4F0", fill_type="solid")
SUBHEADER_FONT = Font(name="微软雅黑", bold=True, size=10)
ALT_ROW_FILL = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
TOTAL_FONT = Font(name="微软雅黑", bold=True, size=10)
TOTAL_BORDER_TOP = Border(top=Side(style="medium", color="1F3864"))
BODY_FONT = Font(name="微软雅黑", size=10)
TITLE_FONT = Font(name="微软雅黑", bold=True, size=14, color="1F3864")
MONEY_FMT = '#,##0'
PCT_FMT = '0.0%'
CENTER = Alignment(horizontal="center", vertical="center")
LEFT = Alignment(horizontal="left", vertical="center")
RIGHT = Alignment(horizontal="right", vertical="center")


def _set_col_widths(ws, widths: list[int]):
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w


def _write_header_row(ws, row: int, values: list[str]):
    for col, val in enumerate(values, 1):
        cell = ws.cell(row=row, column=col, value=val)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = CENTER


def _write_data_row(ws, row: int, values: list, is_alt: bool = False, fmt: str | None = None):
    for col, val in enumerate(values, 1):
        cell = ws.cell(row=row, column=col, value=val)
        cell.font = BODY_FONT
        cell.alignment = RIGHT if isinstance(val, (int, float)) else LEFT
        if is_alt:
            cell.fill = ALT_ROW_FILL
        if fmt and isinstance(val, (int, float)):
            cell.number_format = fmt


def _write_total_row(ws, row: int, values: list, fmt: str | None = None):
    for col, val in enumerate(values, 1):
        cell = ws.cell(row=row, column=col, value=val)
        cell.font = TOTAL_FONT
        cell.border = TOTAL_BORDER_TOP
        cell.alignment = RIGHT if isinstance(val, (int, float)) else LEFT
        if fmt and isinstance(val, (int, float)):
            cell.number_format = fmt


async def generate_quotation_excel(data: dict) -> bytes:
    """Generate a formatted quotation Excel file.

    Args:
        data: Quotation data dict containing cost_breakdown,
              financial_indicators, pricing, and project info.

    Returns:
        Excel file as bytes.
    """
    wb = Workbook()

    _build_summary_sheet(wb, data)
    _build_cost_detail_sheet(wb, data)
    _build_labor_detail_sheet(wb, data)
    _build_pricing_sheet(wb, data)

    buffer = io.BytesIO()
    wb.save(buffer)
    return buffer.getvalue()


# ──────────────────────────────────────────
# Sheet 1: 报价汇总
# ──────────────────────────────────────────

def _build_summary_sheet(wb: Workbook, data: dict):
    ws = wb.active
    ws.title = "报价汇总"
    _set_col_widths(ws, [22, 25, 20, 20])
    ws.sheet_properties.tabColor = "1F3864"

    # Title
    ws.merge_cells("A1:D1")
    title_cell = ws.cell(row=1, column=1, value="物流项目报价汇总")
    title_cell.font = TITLE_FONT
    title_cell.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 36

    # Project info
    info_rows = [
        ("项目名称", data.get("project_name", "")),
        ("客户名称", data.get("client_name", "")),
        ("方案名称", data.get("scheme_name", "")),
        ("报价日期", data.get("date", datetime.now().strftime("%Y-%m-%d"))),
    ]
    for i, (label, value) in enumerate(info_rows, 3):
        ws.cell(row=i, column=1, value=label).font = SUBHEADER_FONT
        ws.cell(row=i, column=2, value=value).font = BODY_FONT

    # Financial indicators
    row = 8
    ws.merge_cells(f"A{row}:D{row}")
    ind_title = ws.cell(row=row, column=1, value="核心财务指标")
    ind_title.font = Font(name="微软雅黑", bold=True, size=12, color="1F3864")
    ind_title.fill = SUBHEADER_FILL

    indicators = data.get("financial_indicators", {})
    ind_items = [
        ("投资回报率 (ROI)", indicators.get("roi_percent"), "%"),
        ("内部收益率 (IRR)", indicators.get("irr_percent"), "%"),
        ("净现值 (NPV, 8%折现)", indicators.get("npv_at_8pct"), "元"),
        ("回本周期", indicators.get("payback_months"), "个月"),
    ]
    for i, (label, value, unit) in enumerate(ind_items, row + 1):
        ws.cell(row=i, column=1, value=label).font = SUBHEADER_FONT
        val_cell = ws.cell(row=i, column=2, value=value)
        val_cell.font = Font(name="微软雅黑", bold=True, size=12, color="2E75B6")
        val_cell.number_format = MONEY_FMT if unit == "元" else "0.0"
        ws.cell(row=i, column=3, value=unit).font = BODY_FONT

    # Pricing summary
    row = 14
    ws.merge_cells(f"A{row}:D{row}")
    pr_title = ws.cell(row=row, column=1, value="报价方案")
    pr_title.font = Font(name="微软雅黑", bold=True, size=12, color="1F3864")
    pr_title.fill = SUBHEADER_FILL

    pricing = data.get("pricing", {})
    pr_items = [
        ("每单价格", pricing.get("per_order"), "元/单"),
        ("每托价格", pricing.get("per_pallet"), "元/托"),
        ("月租单价", pricing.get("per_sqm_month"), "元/㎡/月"),
        ("年度总报价", pricing.get("total_annual"), "元"),
    ]
    for i, (label, value, unit) in enumerate(pr_items, row + 1):
        ws.cell(row=i, column=1, value=label).font = SUBHEADER_FONT
        val_cell = ws.cell(row=i, column=2, value=value)
        val_cell.font = Font(name="微软雅黑", bold=True, size=11)
        val_cell.number_format = MONEY_FMT
        ws.cell(row=i, column=3, value=unit).font = BODY_FONT

    ws.freeze_panes = "A2"


# ──────────────────────────────────────────
# Sheet 2: 成本明细
# ──────────────────────────────────────────

def _build_cost_detail_sheet(wb: Workbook, data: dict):
    ws = wb.create_sheet("成本明细")
    _set_col_widths(ws, [18, 16, 16, 16])
    ws.sheet_properties.tabColor = "2E75B6"

    _write_header_row(ws, 1, ["成本类别", "第1年", "第2年", "第3年"])

    breakdown = data.get("cost_breakdown", {})
    category_names = {
        "labor": "人力成本",
        "facility": "场地成本",
        "equipment": "设备成本",
        "technology": "技术成本",
        "operations": "运营成本",
    }

    row = 2
    totals = [0, 0, 0]
    for key, label in category_names.items():
        cat = breakdown.get(key, {})
        y1 = cat.get("year1", 0) or 0
        y2 = cat.get("year2", 0) or 0
        y3 = cat.get("year3", 0) or 0
        is_alt = (row % 2 == 0)
        _write_data_row(ws, row, [label, y1, y2, y3], is_alt=is_alt, fmt=MONEY_FMT)
        totals[0] += y1
        totals[1] += y2
        totals[2] += y3
        row += 1

    _write_total_row(ws, row, ["合计", *totals], fmt=MONEY_FMT)
    ws.freeze_panes = "A2"


# ──────────────────────────────────────────
# Sheet 3: 人力明细
# ──────────────────────────────────────────

def _build_labor_detail_sheet(wb: Workbook, data: dict):
    ws = wb.create_sheet("人力明细")
    _set_col_widths(ws, [16, 12, 14, 16])
    ws.sheet_properties.tabColor = "548235"

    _write_header_row(ws, 1, ["岗位", "人数", "月均成本", "年成本"])

    labor = data.get("cost_breakdown", {}).get("labor", {})
    details = labor.get("details", [])

    row = 2
    total_count = 0
    total_annual = 0
    for item in details:
        name = item.get("item", "")
        count = item.get("count", 0) or 0
        unit_cost = item.get("unit_cost", 0) or 0
        annual = item.get("annual", 0) or count * unit_cost * 12
        is_alt = (row % 2 == 0)
        _write_data_row(ws, row, [name, count, unit_cost, annual], is_alt=is_alt, fmt=MONEY_FMT)
        total_count += count
        total_annual += annual
        row += 1

    _write_total_row(ws, row, ["合计", total_count, "", total_annual], fmt=MONEY_FMT)
    ws.freeze_panes = "A2"


# ──────────────────────────────────────────
# Sheet 4: 价格方案
# ──────────────────────────────────────────

def _build_pricing_sheet(wb: Workbook, data: dict):
    ws = wb.create_sheet("价格方案")
    _set_col_widths(ws, [20, 16, 14])
    ws.sheet_properties.tabColor = "BF8F00"

    _write_header_row(ws, 1, ["计价项目", "单价", "单位"])

    pricing = data.get("pricing", {})
    items = [
        ("操作费 (每单)", pricing.get("per_order", 0), "元/单"),
        ("仓储费 (每托)", pricing.get("per_pallet", 0), "元/托/天"),
        ("场地费 (每平米)", pricing.get("per_sqm_month", 0), "元/㎡/月"),
    ]

    for i, (label, price, unit) in enumerate(items, 2):
        is_alt = (i % 2 == 0)
        _write_data_row(ws, i, [label, price, unit], is_alt=is_alt, fmt=MONEY_FMT)

    row = len(items) + 3
    ws.merge_cells(f"A{row}:C{row}")
    ann_title = ws.cell(row=row, column=1, value="年度报价汇总")
    ann_title.font = SUBHEADER_FONT
    ann_title.fill = SUBHEADER_FILL

    total = pricing.get("total_annual", 0)
    ws.cell(row=row + 1, column=1, value="年度总报价").font = SUBHEADER_FONT
    total_cell = ws.cell(row=row + 1, column=2, value=total)
    total_cell.font = Font(name="微软雅黑", bold=True, size=14, color="C00000")
    total_cell.number_format = MONEY_FMT
    ws.cell(row=row + 1, column=3, value="元/年").font = BODY_FONT

    ws.freeze_panes = "A2"
